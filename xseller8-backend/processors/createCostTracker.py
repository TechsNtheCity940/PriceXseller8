# Let's now apply formatting to make the document look official
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import openpyxl

# Create a new Excel workbook and select the active worksheet

wb = openpyxl.Workbook()

ws = wb.active



# Rename the active sheet to a meaningful name (optional)

ws.title = "Cost Tracker"



# Define the headers based on the provided columns

headers = ["Date", "Product number", "Product", "Brand", "Pkg/size", 

           "Price (per \".\")", "Ordered", "Confirmed", "Status"]



# Start adding headers from column B, row 2

for col_num, header in enumerate(headers, start=2):

    ws.cell(row=2, column=col_num).value = header


# Define the border style for the cells
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Apply formatting to the header row (B2 onwards for headers)
for col_num in range(2, 2 + len(headers)):
    cell = ws.cell(row=2, column=col_num)
    
    # Apply border
    cell.border = thin_border
    
    # Apply font styling (bold and slightly larger font size)
    cell.font = Font(bold=True, size=12)
    
    # Center alignment
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply fill color for header (light grey background)
    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

# Apply borders and center alignment for the entire table area (15 columns, 35 rows, starting from B2)
for row in range(2, 37):  # Row 2 to Row 36
    for col in range(2, 17):  # Column B to Column P
        cell = ws.cell(row=row, column=col)
        
        # Apply border
        cell.border = thin_border
        
        # Center alignment for all cells
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the workbook with the new formatting
file_path = 'D:/GitRepos/PriceXseller8/xseller8-backend/documents/cost_tracker.xlsx'
wb.save(file_path)

file_path  # Returning the formatted file path
